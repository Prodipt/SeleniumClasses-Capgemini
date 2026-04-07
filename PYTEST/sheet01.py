import pytest
# from openpyxl.workbook import Workbook

import openpyxl

wb = openpyxl.Workbook()
sheetName = "sheet1"
if sheetName in wb.sheetnames:
    ws = wb[sheetName]          #ws : worksheet
else:
    ws = wb.create_sheet(sheetName)

ws['A1'] = 'USER'
ws['B1'] = 'PASSWORD'
ws['C1'] = 'RANK'

# wb.save('sample.xlsx')   # It will save in the local directory

# Append used to add data in the worksheet
ws.append(['standard_user', 'secret_sauce', 'Products'] )
ws.append(['standard_user', 'secret_sauce', 'Products'] )
# ws.append(['wrong_user', 'wrong_password', 'Products'] )



# wb.save('sample.xlsx')   #To save the file


# ws.delete_rows(2)

for row in ws.iter_rows(values_only=True): #To Iterate in the Excel Sheet
    print(row)

wb.save('sample.xlsx')   #To save the file
